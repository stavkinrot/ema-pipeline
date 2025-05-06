# --- outlier_detector.py ---
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from processing.contradiction_config import CONTRADICTION_PAIRS, COMPATIBLE_PAIRS
from openpyxl.comments import Comment
from datetime import timedelta


class OutlierDetector:
    def __init__(self, df: pd.DataFrame, id_col='Participant ID', code_col='participant_code'):
        self.df = df.copy()
        self.id_col = id_col
        self.code_col = code_col
        self.results = defaultdict(lambda: defaultdict(list))
        self.flagged_cells = defaultdict(list)
        self.HIGH_THRESHOLD = 4
        self.LOW_THRESHOLD = 4

    def _safe_int(self, val):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None

    def detect_constant_answers(self):
        for idx, row in self.df.iterrows():
            pid = row[self.code_col]
            answers = row.drop(labels=[self.id_col, self.code_col, 'day', 'time_of_day', 'first_day'])
            ints = answers.map(self._safe_int).dropna()
            if len(ints) >= 14 and ints.nunique() == 1:
                self.results[pid][idx].append('constant')

    def detect_middle_only(self):
        for idx, row in self.df.iterrows():
            pid = row[self.code_col]
            answers = row.drop(labels=[self.id_col, self.code_col, 'day', 'time_of_day', 'first_day'])
            ints = answers.map(self._safe_int).dropna()
            if len(ints) >= 14 and all(a == 4 for a in ints):
                self.results[pid][idx].append('middle_only')

    def detect_edge_only(self):
        for idx, row in self.df.iterrows():
            pid = row[self.code_col]
            answers = row.drop(labels=[self.id_col, self.code_col, 'day', 'time_of_day', 'first_day'])
            ints = answers.map(self._safe_int).dropna()
            if len(ints) >= 16 and all(a in {1, 7} for a in ints):
                self.results[pid][idx].append('edge_only')

    def detect_contradictions(self):
        for idx, row in self.df.iterrows():
            pid = row[self.code_col]
            for q1, q2 in CONTRADICTION_PAIRS:
                v1, v2 = self._safe_int(row.get(q1)), self._safe_int(row.get(q2))
                if v1 and v1 > self.HIGH_THRESHOLD and v2 and v2 > self.HIGH_THRESHOLD:
                    self.results[pid][idx].append('contradiction')
                    self.flagged_cells[(idx, q1)] = 'contradiction'
                    self.flagged_cells[(idx, q2)] = 'contradiction'

    def detect_incompatibilities(self):
        for idx, row in self.df.iterrows():
            pid = row[self.code_col]
            for q1, q2 in COMPATIBLE_PAIRS:
                v1, v2 = self._safe_int(row.get(q1)), self._safe_int(row.get(q2))
                if v1 and v2:
                    if (v1 > self.HIGH_THRESHOLD and v2 < self.LOW_THRESHOLD) or \
                       (v1 < self.LOW_THRESHOLD and v2 > self.HIGH_THRESHOLD):
                        self.results[pid][idx].append('compatible_mismatch')
                        self.flagged_cells[(idx, q1)] = 'compatible_mismatch'
                        self.flagged_cells[(idx, q2)] = 'compatible_mismatch'

    def detect_time_mismatches(self):
        print("[DEBUG] Running parent-child time mismatch detection")
        rows_by_key = {}
        for idx, row in self.df.iterrows():
            key = (row[self.code_col], row['day'], row['time_of_day'])
            rows_by_key[key] = (idx, row)

        for (code, day, tod), (idx, child_row) in rows_by_key.items():
            try:
                if int(code[1:]) % 2 != 0:
                    continue  # skip parents
                parent_code = f"#{int(code[1:]) + 1:04}"
                parent_key = (parent_code, day, tod)
                if parent_key not in rows_by_key:
                    continue
                parent_idx, parent_row = rows_by_key[parent_key]

                # Force timestamp parsing
                print("IM HERE")
                child_ts = pd.to_datetime(child_row.get("timestamp"), errors="coerce")
                parent_ts = pd.to_datetime(parent_row.get("timestamp"), errors="coerce")

                if pd.isna(child_ts) or pd.isna(parent_ts):
                    print(f"[DEBUG] Missing timestamp(s) for day {day}, {tod} between {code} and {parent_code}")
                    continue

                delta = abs(parent_ts - child_ts)
                print(f"[DEBUG] Time difference on day {day}, {tod} between {code} and {parent_code}: {delta}")
                if delta > timedelta(minutes=15):
                    self.flagged_cells[(idx, "timestamp")] = 'time_mismatch'
                    self.flagged_cells[(parent_idx, "timestamp")] = 'time_mismatch'
            except Exception as e:
                print(f"[DEBUG] Exception in mismatch detection for {code}: {e}")
                continue


    def detect_outliers(self):
        self.detect_constant_answers()
        self.detect_middle_only()
        self.detect_edge_only()
        self.detect_contradictions()
        self.detect_incompatibilities()
        self.detect_time_mismatches()

    def highlight_in_excel(self, xlsx_path_in, xlsx_path_out):
        wb = load_workbook(xlsx_path_in)
        ws = wb.active

        color_map = {
            'constant': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
            'middle_only': PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid'),
            'edge_only': PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid'),
            'contradiction': PatternFill(start_color='FF99CC', end_color='FF99CC', fill_type='solid'),
            'compatible_mismatch': PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'),
            'very_suspicious': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'time_mismatch': PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
        }

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        id_idx = header.index(self.id_col)
        code_idx = header.index(self.code_col)
        col_map = {col: i for i, col in enumerate(header)}

        survey_counts = defaultdict(int)
        for pid, row_flags in self.results.items():
            survey_counts[pid] = len(row_flags)

        for pid, row_flags in self.results.items():
            for idx, flags in row_flags.items():
                row_idx = idx + 2
                if survey_counts[pid] > 7:
                    for col_idx in [id_idx, code_idx]:
                        cell = ws.cell(row=row_idx, column=col_idx + 1)
                        cell.fill = color_map['very_suspicious']
                        cell.comment = Comment("very_suspicious", "OutlierDetector")

                non_specific_flags = [f for f in flags if f not in ['contradiction', 'compatible_mismatch']]
                if non_specific_flags:
                    main_flag = non_specific_flags[0]
                    for cell in ws[row_idx]:
                        cell.fill = color_map[main_flag]
                        cell.comment = Comment(main_flag, "OutlierDetector")

        for (row_idx, col), reason in self.flagged_cells.items():
            if col in col_map:
                col_idx = col_map[col]
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                cell.fill = color_map[reason]
                cell.comment = Comment(reason, "OutlierDetector")

        legend_start_row = ws.max_row + 3
        ws.cell(row=legend_start_row, column=1, value="Legend:")
        for i, (flag, fill) in enumerate(color_map.items()):
            cell = ws.cell(row=legend_start_row + i + 1, column=1, value=flag)
            cell.fill = fill

        wb.save(xlsx_path_out)
